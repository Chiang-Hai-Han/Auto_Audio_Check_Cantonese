"""独立预检查/预整理模块。"""
from __future__ import annotations

import datetime as _dt
import json
import logging
import os
from dataclasses import dataclass
from typing import Callable, Dict, List, Optional

from openpyxl import Workbook

from matcher import ExcelMatcher
from scanner import AudioScanner

logger = logging.getLogger(__name__)

ProgressCallback = Optional[Callable[[float, str], None]]


@dataclass
class PreflightResult:
    input_dir: str
    excel_files: List[str]
    prepared_excel_path: str
    preview_report_path: str
    total_audio: int
    matched_audio: int
    unmatched_audio: int
    filename_parse_failed: int
    low_confidence_count: int
    severity: str
    blockers: List[str]
    warnings: List[str]
    rows: List[Dict[str, object]]
    prepare_summary: Dict[str, object]


def _emit(progress_cb: ProgressCallback, value: float, message: str) -> None:
    if progress_cb:
        progress_cb(value, message)


def load_config(config_path: str = "config.json") -> dict:
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    return {}


def _report_dir(config: dict) -> str:
    return config.get("output", {}).get("report_dir", "reports")


def _timestamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _export_prepared_excel(report_dir: str, sheets_data: Dict[str, List[dict]], folder_name: str) -> str:
    os.makedirs(report_dir, exist_ok=True)
    path = os.path.join(report_dir, f"预整理配音表_{folder_name}_{_timestamp()}.xlsx")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=(sheet_name or "Sheet")[:31])
        ws.append(["角色", "编号", "台词", "语速", "语气", "来源文件", "原Sheet", "原行号"])
        for row in rows:
            ws.append([
                row.get("role", ""),
                row.get("role_index", ""),
                row.get("text", ""),
                row.get("speed", ""),
                row.get("tone", ""),
                os.path.basename(row.get("source_file", "")),
                row.get("original_sheet", ""),
                row.get("row", ""),
            ])

    wb.save(path)
    return path


def _export_preview_report(report_dir: str, rows: List[Dict[str, object]], folder_name: str) -> str:
    os.makedirs(report_dir, exist_ok=True)
    path = os.path.join(report_dir, f"检查预览_{folder_name}_{_timestamp()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "匹配预览"
    ws.append([
        "文件夹", "音频文件", "角色", "序号", "匹配分类",
        "预期台词", "匹配状态", "置信度", "来源文件", "来源Sheet", "来源行号"
    ])

    for row in rows:
        ws.append([
            row["folder_name"],
            row["filename"],
            row["role"],
            row["role_index"],
            row["sheet_name"],
            row["expected_text"],
            row["match_status"],
            row["match_score"],
            row["source_file"],
            row["source_sheet"],
            row["source_row"],
        ])

    wb.save(path)
    return path


def _build_rows(audios: list[object], sheets_data: Dict[str, List[dict]]) -> List[Dict[str, object]]:
    source_lookup = {}
    for rows in sheets_data.values():
        for row in rows:
            key = (
                row.get("sheet", ""),
                row.get("text", ""),
                row.get("role", ""),
                row.get("role_index", ""),
                row.get("row", 0),
            )
            source_lookup[key] = row

    preview_rows: List[Dict[str, object]] = []
    for meta in audios:
        key = (
            getattr(meta, "sheet_name", ""),
            getattr(meta, "expected_text_from_excel", ""),
            getattr(meta, "role", ""),
            getattr(meta, "role_index", ""),
            getattr(meta, "row_number", 0),
        )
        source = source_lookup.get(key, {})
        expected_text = getattr(meta, "expected_text_from_excel", "") or getattr(meta, "expected_text_from_filename", "")
        preview_rows.append({
            "folder_name": getattr(meta, "folder_name", ""),
            "filename": getattr(meta, "filename", ""),
            "role": getattr(meta, "role", ""),
            "role_index": getattr(meta, "role_index", ""),
            "sheet_name": getattr(meta, "sheet_name", ""),
            "expected_text": expected_text,
            "match_status": "已匹配" if getattr(meta, "expected_text_from_excel", "") else "未匹配",
            "match_score": getattr(meta, "match_score", 0.0),
            "source_file": os.path.basename(source.get("source_file", "")),
            "source_sheet": source.get("original_sheet", ""),
            "source_row": source.get("row", ""),
        })
    return preview_rows


def _judge_severity(
    total_audio: int,
    matched_audio: int,
    filename_parse_failed: int,
    low_confidence_count: int,
    excel_files: List[str],
) -> tuple[str, List[str], List[str]]:
    blockers: List[str] = []
    warnings: List[str] = []

    if total_audio == 0:
        blockers.append("没有扫描到任何 wav 音频文件。")
    if not excel_files:
        blockers.append("没有找到可用的 Excel 文件。")

    match_rate = (matched_audio / total_audio) if total_audio else 0.0
    parse_fail_rate = (filename_parse_failed / total_audio) if total_audio else 0.0

    if total_audio and match_rate < 0.8:
        blockers.append(f"匹配成功率过低：{matched_audio}/{total_audio}。")
    elif total_audio and match_rate < 0.95:
        warnings.append(f"匹配成功率偏低：{matched_audio}/{total_audio}。")

    if parse_fail_rate > 0.2:
        blockers.append(f"文件名解析失败过多：{filename_parse_failed}/{total_audio}。")
    elif filename_parse_failed:
        warnings.append(f"存在文件名解析失败：{filename_parse_failed} 条。")

    if low_confidence_count > 0:
        warnings.append(f"存在低置信度匹配：{low_confidence_count} 条。")

    if blockers:
        return "block", blockers, warnings
    if warnings:
        return "warn", blockers, warnings
    return "pass", blockers, warnings


def run_preflight(input_dir: str, config_path: str = "config.json", progress_cb: ProgressCallback = None) -> PreflightResult:
    config = load_config(config_path)
    report_dir = _report_dir(config)
    folder_name = os.path.basename(input_dir.rstrip(os.sep))

    _emit(progress_cb, 0.05, "正在扫描音频文件")
    scanner = AudioScanner()
    audios = scanner.scan(input_dir)

    _emit(progress_cb, 0.2, "正在查找 Excel 文件")
    matcher = ExcelMatcher(config)
    excel_path = matcher.find_excel_in_dir(input_dir)
    excel_files = matcher.find_excel_files(input_dir)

    _emit(progress_cb, 0.45, "正在预整理 Excel")
    sheets_data = matcher.load_excel(excel_path) if excel_path else {}

    _emit(progress_cb, 0.7, "正在干跑匹配")
    audios = matcher.match_audios(audios, sheets_data) if sheets_data else audios
    rows = _build_rows(audios, sheets_data)

    _emit(progress_cb, 0.85, "正在导出预整理结果")
    prepared_excel_path = _export_prepared_excel(report_dir, sheets_data, folder_name) if sheets_data else ""
    preview_report_path = _export_preview_report(report_dir, rows, folder_name) if rows else ""

    total_audio = len(audios)
    matched_audio = sum(1 for audio in audios if getattr(audio, "expected_text_from_excel", ""))
    unmatched_audio = total_audio - matched_audio
    filename_parse_failed = sum(1 for audio in audios if not getattr(audio, "role", ""))
    low_confidence_count = sum(
        1 for audio in audios
        if getattr(audio, "expected_text_from_excel", "") and getattr(audio, "match_score", 1.0) < 0.75
    )

    severity, blockers, warnings = _judge_severity(
        total_audio=total_audio,
        matched_audio=matched_audio,
        filename_parse_failed=filename_parse_failed,
        low_confidence_count=low_confidence_count,
        excel_files=excel_files,
    )

    _emit(progress_cb, 1.0, "检查完成")
    return PreflightResult(
        input_dir=input_dir,
        excel_files=excel_files,
        prepared_excel_path=prepared_excel_path,
        preview_report_path=preview_report_path,
        total_audio=total_audio,
        matched_audio=matched_audio,
        unmatched_audio=unmatched_audio,
        filename_parse_failed=filename_parse_failed,
        low_confidence_count=low_confidence_count,
        severity=severity,
        blockers=blockers,
        warnings=warnings,
        rows=rows,
        prepare_summary=dict(matcher.last_prepare_summary),
    )

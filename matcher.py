"""Load split Excel files, normalize rows, and match audios."""
import logging
import os
import re
import shutil
import tempfile
from difflib import SequenceMatcher
from typing import Dict, List, Tuple

import openpyxl

from ai_engine.base import AudioMeta

logger = logging.getLogger(__name__)


class ExcelMatcher:
    FOLDER_SHEET_MAP = {
        "剧情动画_配音": "剧情动画",
        "剧情动画配音": "剧情动画",
        "剧情_配音": "剧情动画",
        "预习_配音": "预习配音",
        "预习配音": "预习配音",
        "线上作业_配音": "线上作业",
        "作业配音": "线上作业",
        "总结动画_配音": "总结动画",
        "总结语音": "总结动画",
        "题目_配音": "题目配音",
        "题目语音": "题目配音",
        "系统语音": "题目配音",
        "Sheet1": "Sheet1",
    }

    EXCEL_IGNORE_KEYWORDS = ("配音合集",)

    def __init__(self, config: dict):
        self.config = config or {}
        self.last_prepare_summary: Dict[str, object] = {}

    def find_excel_files(self, root_dir: str) -> List[str]:
        candidates = []
        fallback = []
        for dirpath, _, filenames in os.walk(root_dir):
            for fname in filenames:
                if fname.startswith("~$"):
                    continue
                if not fname.lower().endswith((".xlsx", ".xls")):
                    continue
                path = os.path.join(dirpath, fname)
                if any(keyword in fname for keyword in self.EXCEL_IGNORE_KEYWORDS):
                    fallback.append(path)
                else:
                    candidates.append(path)

        selected = sorted(candidates) if candidates else sorted(fallback)
        if selected:
            logger.info("发现 Excel 文件: %s", " | ".join(os.path.basename(p) for p in selected))
        return selected

    def find_excel_in_dir(self, root_dir: str) -> str:
        files = self.find_excel_files(root_dir)
        if not files:
            return ""
        if len(files) > 1:
            logger.info("发现 %s 个拆分 Excel，将自动汇总后使用", len(files))
        else:
            logger.info("选择 Excel: %s", files[0])
        return files[0]

    def _open_workbook(self, excel_path: str):
        try:
            return openpyxl.load_workbook(excel_path, data_only=True)
        except OSError:
            tmp = os.path.join(tempfile.gettempdir(), "_audit_tmp.xlsx")
            shutil.copy2(excel_path, tmp)
            wb = openpyxl.load_workbook(tmp, data_only=True)
            os.remove(tmp)
            return wb

    def _detect_columns(self, ws) -> Dict[str, int]:
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {"role": -1, "index": -1, "text": -1, "speed": -1, "tone": -1}

        role_keywords = ["角色", "人物", "名称", "名字"]
        index_keywords = ["编号", "序号", "id"]
        text_keywords = ["对白", "台词", "旁白", "内容", "字幕", "文案"]
        speed_keywords = ["语速", "速度"]
        tone_keywords = ["语气", "情绪", "情感"]

        for i, val in enumerate(header_row):
            if val is None:
                continue
            s = str(val).strip().lower()
            if col_map["role"] < 0 and any(k in s for k in role_keywords):
                col_map["role"] = i
            elif col_map["index"] < 0 and any(k in s for k in index_keywords):
                col_map["index"] = i
            elif col_map["text"] < 0 and any(k in s for k in text_keywords):
                col_map["text"] = i
            elif col_map["speed"] < 0 and any(k in s for k in speed_keywords):
                col_map["speed"] = i
            elif col_map["tone"] < 0 and any(k in s for k in tone_keywords):
                col_map["tone"] = i

        if col_map["role"] < 0:
            col_map["role"] = 0
        if col_map["text"] < 0:
            col_map["text"] = 3 if len(header_row) > 3 else max(len(header_row) - 1, 0)
        if col_map["speed"] < 0:
            col_map["speed"] = 5 if len(header_row) > 5 else -1
        if col_map["tone"] < 0:
            col_map["tone"] = 4 if len(header_row) > 4 else -1

        logger.info(
            "  列检测: role=%s, idx=%s, text=%s, speed=%s, tone=%s",
            col_map["role"], col_map["index"], col_map["text"], col_map["speed"], col_map["tone"]
        )
        return col_map

    @staticmethod
    def _normalize_text(text: str) -> str:
        if not text:
            return ""
        text = str(text).strip().replace("\n", "").replace("\r", "")
        text = re.sub(r"\s+", "", text)
        text = text.replace("：", "").replace(":", "")
        text = text.replace("……", "…")
        text = text.replace("～", "").replace("~", "")
        text = text.replace("一一", "一")
        return text

    @staticmethod
    def _normalize_role(role: str, role_index: str = "") -> Tuple[str, str]:
        role = str(role or "").strip()
        role = re.sub(r"[（(]\s*(\d+)\s*[)）]", r"\1", role)
        role = role.replace(" ", "")
        idx = str(role_index or "").strip()

        m = re.match(r"^(?P<name>.*?)(?P<num>\d+)$", role)
        if m and not idx:
            role = m.group("name")
            idx = m.group("num")

        if idx.isdigit():
            idx = idx.zfill(2)

        return role, idx

    def _infer_sheet_category(self, source_path: str, sheet_name: str) -> str:
        source_name = os.path.basename(source_path)
        combined = f"{source_name} {sheet_name}"
        for keyword, target in self.FOLDER_SHEET_MAP.items():
            if keyword in combined:
                return target
        return sheet_name.strip() or os.path.splitext(source_name)[0]

    def load_excel(self, excel_path: str) -> Dict[str, List[dict]]:
        root_dir = os.path.dirname(excel_path)
        excel_files = self.find_excel_files(root_dir)
        if excel_path not in excel_files and os.path.exists(excel_path):
            excel_files.insert(0, excel_path)

        sheets_data: Dict[str, List[dict]] = {}
        source_count = 0
        row_count = 0
        empty_text_rows = 0
        duplicate_rows = 0
        seen_keys = set()

        for path in excel_files:
            try:
                wb = self._open_workbook(path)
            except Exception as exc:
                logger.warning("读取 Excel 失败，跳过: %s | %s", path, exc)
                continue

            source_count += 1
            logger.info("载入 Excel: %s | %s 个 Sheet", path, len(wb.sheetnames))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                col_map = self._detect_columns(ws)
                category = self._infer_sheet_category(path, sheet_name)
                rows = sheets_data.setdefault(category, [])

                for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    role_val = ""
                    if col_map["role"] >= 0 and len(row) > col_map["role"] and row[col_map["role"]]:
                        role_val = str(row[col_map["role"]]).strip()

                    idx_val = ""
                    if col_map["index"] >= 0 and len(row) > col_map["index"] and row[col_map["index"]]:
                        idx_val = str(row[col_map["index"]]).strip()

                    text_val = ""
                    if col_map["text"] >= 0 and len(row) > col_map["text"] and row[col_map["text"]]:
                        text_val = str(row[col_map["text"]]).strip()

                    if not role_val and not text_val:
                        continue
                    if not text_val:
                        empty_text_rows += 1
                        continue

                    role_val, idx_val = self._normalize_role(role_val, idx_val)
                    norm_text = self._normalize_text(text_val)

                    speed_val = "常规"
                    if col_map["speed"] >= 0 and len(row) > col_map["speed"] and row[col_map["speed"]]:
                        speed_val = str(row[col_map["speed"]]).strip()

                    tone_val = "常规"
                    if col_map["tone"] >= 0 and len(row) > col_map["tone"] and row[col_map["tone"]]:
                        tone_val = str(row[col_map["tone"]]).strip()

                    dedupe_key = (category, role_val, idx_val, norm_text)
                    if dedupe_key in seen_keys:
                        duplicate_rows += 1
                        continue
                    seen_keys.add(dedupe_key)

                    rows.append({
                        "sheet": category,
                        "original_sheet": sheet_name,
                        "source_file": path,
                        "row": row_number,
                        "role": role_val,
                        "role_index": idx_val,
                        "speed": speed_val,
                        "tone": tone_val,
                        "text": text_val.strip().replace("\n", ""),
                        "norm_text": norm_text,
                    })
                    row_count += 1

            wb.close()

        self.last_prepare_summary = {
            "excel_files": excel_files,
            "loaded_files": source_count,
            "sheet_count": len(sheets_data),
            "row_count": row_count,
            "empty_text_rows": empty_text_rows,
            "duplicate_rows": duplicate_rows,
        }
        logger.info(
            "Excel 预整理完成：%s 份文件，%s 个分类，%s 条有效台词",
            source_count, len(sheets_data), row_count
        )
        return sheets_data

    def match_folder_to_sheet(self, folder_name: str, sheet_names: List[str]) -> str:
        if folder_name in sheet_names:
            return folder_name
        for keyword, sheet in self.FOLDER_SHEET_MAP.items():
            if keyword in folder_name and sheet in sheet_names:
                return sheet

        best = ""
        best_score = 0.0
        for sheet_name in sheet_names:
            score = SequenceMatcher(None, folder_name, sheet_name).ratio()
            if score > best_score:
                best_score = score
                best = sheet_name
        return best if best_score > 0.3 else ""

    def _score_row(self, meta: AudioMeta, row: dict) -> float:
        filename_text = self._normalize_text(meta.expected_text_from_filename)
        row_text = row.get("norm_text") or self._normalize_text(row["text"])

        text_score = SequenceMatcher(None, filename_text, row_text).ratio() if filename_text else 0.0

        role, idx = self._normalize_role(meta.role, meta.role_index)
        row_role, row_idx = self._normalize_role(row["role"], row["role_index"])

        role_score = 0.0
        if role and row_role:
            role_score = 1.0 if role == row_role else SequenceMatcher(None, role, row_role).ratio()
        idx_score = 0.0
        if idx and row_idx:
            idx_score = 1.0 if idx == row_idx else 0.0

        return text_score * 0.75 + role_score * 0.20 + idx_score * 0.05

    def match_audios(self, audios: List[AudioMeta], sheets_data: Dict[str, List[dict]]) -> List[AudioMeta]:
        sheet_names = list(sheets_data.keys())
        matched = 0

        for meta in audios:
            sheet_name = self.match_folder_to_sheet(meta.folder_name, sheet_names)
            meta.sheet_name = sheet_name
            if not sheet_name or sheet_name not in sheets_data:
                continue

            rows = sheets_data[sheet_name]
            best_row = None
            best_score = 0.0
            for row in rows:
                score = self._score_row(meta, row)
                if score > best_score:
                    best_score = score
                    best_row = row

            if best_row and best_score >= 0.55:
                self._apply_match(meta, best_row)
                meta.match_score = round(best_score, 4)
                matched += 1

        logger.info("匹配完成：%s/%s 条音频成功匹配", matched, len(audios))
        return audios

    @staticmethod
    def _apply_match(meta: AudioMeta, row: dict):
        meta.expected_text_from_excel = row["text"]
        meta.expected_speed = row["speed"]
        meta.expected_tone = row["tone"]
        meta.row_number = row["row"]

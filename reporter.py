"""
報告生成器：輸出審核結果 Excel 報告
"""
import os
import datetime
import logging
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ai_engine.base import ReviewResult

logger = logging.getLogger(__name__)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
NORMAL_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


class Reporter:
    HEADERS = [
        "序號", "文件名", "角色", "所屬文件夾",
        "預期台詞(Excel)", "預期語速", "預期語氣",
        "內容匹配", "語調節奏", "發音準確性",
        "整體判斷", "整體結論", "問題摘要", "錯誤信息"
    ]
    COL_WIDTHS = [6, 35, 12, 15, 40, 8, 10, 25, 20, 25, 10, 35, 40, 25]

    def __init__(self, config: dict):
        output_cfg = config.get("output", {})
        self.report_dir = output_cfg.get("report_dir", "reports")
        self.report_prefix = output_cfg.get("report_prefix", "审核报告")

    def generate(self, results: List[ReviewResult], input_dir: str) -> str:
        os.makedirs(self.report_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = os.path.basename(input_dir.rstrip(os.sep))
        report_name = f"{self.report_prefix}_{folder_name}_{timestamp}.xlsx"
        report_path = os.path.join(self.report_dir, report_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "審核明細"
        self._write_headers(ws)

        for i, result in enumerate(results, 1):
            self._write_result_row(ws, i + 1, result, i)

        ws.freeze_panes = "H2"
        for col_idx, width in enumerate(self.COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        self._create_summary_sheet(wb, results, folder_name, timestamp)
        wb.save(report_path)
        logger.info(f"報告已保存: {report_path}")
        return report_path

    def _write_headers(self, ws):
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def _write_result_row(self, ws, row_num: int, result: ReviewResult, seq: int):
        meta = result.audio_meta
        items = [
            result.content_match,
            result.intonation,
            result.pronunciation
        ]

        values = [
            seq, meta.filename,
            f"{meta.role}{meta.role_index}",
            meta.folder_name,
            meta.expected_text_from_excel or meta.expected_text_from_filename,
            meta.expected_speed, meta.expected_tone,
            self._fmt(result.content_match),
            self._fmt(result.intonation),
            self._fmt(result.pronunciation),
            "PASS" if result.overall_pass else "NG",
            result.overall_summary,
            self._issues(items), result.error,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        if result.error:
            fill = RED_FILL
        elif result.all_passed:
            fill = GREEN_FILL
        else:
            fill = YELLOW_FILL
        for col_idx in range(1, len(values) + 1):
            ws.cell(row=row_num, column=col_idx).fill = fill

    def _fmt(self, item) -> str:
        return "OK" if item.passed else f"NG: {item.issue}"

    def _issues(self, items) -> str:
        issues = [it for it in items if not it.passed]
        if not issues:
            return "無"
        return "; ".join(f"[{it.dimension}] {it.issue}" for it in issues)

    def _create_summary_sheet(self, wb, results, folder_name, timestamp):
        ws = wb.create_sheet("統計摘要")
        total = len(results)
        passed = sum(1 for r in results if r.all_passed and not r.error)
        with_issues = sum(1 for r in results if not r.all_passed)
        with_errors = sum(1 for r in results if r.error)

        rows = [
            ("統計摘要", ""), ("", ""),
            ("審核文件夾", folder_name), ("生成時間", timestamp), ("", ""),
            ("總音頻數", str(total)),
            ("完全通過", f"{passed} ({self._pct(passed, total)})"),
            ("有問題", f"{with_issues} ({self._pct(with_issues, total)})"),
            ("處理失敗", f"{with_errors} ({self._pct(with_errors, total)})"),
            ("", ""), ("--- 各維度問題統計 ---", ""),
        ]
        for i, (label, value) in enumerate(rows, 1):
            ws.cell(row=i, column=1, value=label).font = BOLD_FONT if "---" in label or "統計" in label else NORMAL_FONT
            ws.cell(row=i, column=2, value=value).font = NORMAL_FONT

        dims = [
            ("內容匹配", "content_match"), ("語調節奏", "intonation"), ("發音準確性", "pronunciation"),
        ]
        r = len(rows) + 1
        for dim_name, dim_attr in dims:
            fail = sum(1 for r_ in results if not getattr(r_, dim_attr).passed)
            ws.cell(row=r, column=1, value=f"  {dim_name} 不通過").font = NORMAL_FONT
            ws.cell(row=r, column=2, value=f"{fail} ({self._pct(fail, total)})").font = NORMAL_FONT
            r += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _pct(part: int, total: int) -> str:
        if total == 0:
            return "0%"
        return f"{part / total * 100:.1f}%"